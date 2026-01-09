import asyncio
import random
import time
import os
import json
import re
import csv
import html
from datetime import datetime
from playwright.async_api import async_playwright
from bs4 import BeautifulSoup

# ==========================================
# 1. KONFIGURACJA
# ==========================================

USER_DATA_DIR = "./chrome_profile_multi"
# Nazwy plików z dopiskiem _parallel
REPORT_FILE = "raport_targeted_parallel.html" 
HISTORY_FILE = "ceny_targeted_parallel.json"
DEBUG_JSON_FILE = "debug_targeted_parallel.json"
DEBUG_CSV_FILE = "debug_targeted_parallel.csv"
DEBUG_XLSX_FILE = "debug_targeted_parallel.xlsx"
CSV_FILE = "raport_targeted_parallel_excel.csv"
CSV_HISTORY_FILE = "historia_cen_targeted_parallel.csv"

MIN_PRICE_THRESHOLD = 600.00 
MAX_PAGES_PER_TASK = 50 

# ==========================================
# 2. FUNKCJE POMOCNICZE
# ==========================================

async def handle_cookies(page, cookie_texts):
    try:
        for text in cookie_texts:
            btn = page.get_by_role("button", name=re.compile(text, re.IGNORECASE))
            if await btn.count() > 0 and await btn.first.is_visible():
                await btn.first.click()
                await asyncio.sleep(0.5)
                return True
    except Exception: pass
    return False

async def human_scroll(page):
    try:
        last_height = await page.evaluate("document.body.scrollHeight")
        while True:
            await page.mouse.wheel(0, 800)
            await asyncio.sleep(random.uniform(0.3, 0.7))
            current_scroll = await page.evaluate("window.scrollY + window.innerHeight")
            new_height = await page.evaluate("document.body.scrollHeight")
            if current_scroll >= new_height - 100: break
            if new_height > last_height: last_height = new_height
    except Exception: pass
    await asyncio.sleep(1.5)

# ==========================================
# 3. PARSERY SKLEPÓW
# ==========================================

def parse_xkom(soup, module_type):
    products = []
    candidates = soup.find_all('div', attrs={'data-name': 'productCard'})
    candidates = list(set(candidates))
    for div in candidates:
        try:
            if "sprawdź też" in div.text.lower(): continue
            title = div.find('h3').text.strip()
            price_match = re.search(r'(\d[\d\s]*[,.]\d{2})\s*zł', div.text)
            if not price_match: continue
            price = float(re.sub(r'[^\d,]', '', price_match.group(1)).replace(',', '.'))
            link_tag = div.find('a')
            link = "https://www.x-kom.pl" + link_tag['href'] if link_tag else "#"
            img = div.find('img')['src'] if div.find('img') else ""
            if price >= MIN_PRICE_THRESHOLD:
                products.append({ 'title': title, 'price': price, 'link': link, 'image': img, 'shop': 'x-kom', 'type': module_type })
        except Exception: continue
    return products

def parse_morele(soup, module_type):
    products = []
    candidates = soup.find_all('div', attrs={'data-product-id': True})
    for div in candidates:
        try:
            title_tag = div.find('a', class_='productLink') or div.find('a', title=True)
            if not title_tag: continue
            title = title_tag.get('title', title_tag.text).strip()
            full_text = div.get_text()
            price_tag = div.find(class_='price-new')
            price_text = price_tag.get_text() if price_tag else full_text
            if "rata" in div.text.lower() and not price_tag: continue 
            price_match = re.search(r'(\d[\d\s]*[,.]?\d{0,2})\s*zł', price_text)
            if not price_match: continue
            price = float(re.sub(r'[^\d,]', '', price_match.group(1)).replace(',', '.'))
            link = title_tag['href']
            link = "https://www.morele.net" + link if link.startswith('/') else link
            img_tag = div.find('img')
            img = img_tag.get('data-src') or img_tag.get('src') or ""
            if price >= MIN_PRICE_THRESHOLD:
                products.append({ 'title': title, 'price': price, 'link': link, 'image': img, 'shop': 'morele', 'type': module_type })
        except Exception: continue
    return products

def parse_sferis(soup, module_type):
    products = []
    price_pattern = re.compile(r'(\d[\d\s\.]*[,]\d{2})\s*(?:z|Z|PLN)?')
    for element in soup.find_all(string=True):
        if element.parent.name in ['script', 'style', 'head']: continue
        text = element.strip()
        if not text: continue
        match = price_pattern.search(text)
        if match:
            try:
                price_str = match.group(1)
                price = float(price_str.replace(" ", "").replace("\xa0", "").replace(".", "").replace(",", "."))
                if price < MIN_PRICE_THRESHOLD: continue
                container = element.parent
                title = ""
                link = ""
                img = ""
                for _ in range(6): 
                    if not container: break
                    if not link:
                        link_tag = container.find('a', href=True)
                        if link_tag:
                            h = link_tag['href']
                            if len(h) > 5 and "javascript" not in h:
                                t = link_tag.get_text(strip=True)
                                if len(t) > 10: 
                                    title = t
                                    link = h
                                elif link_tag.has_attr('title'):
                                    title = link_tag['title']
                                    link = h
                    if not img:
                        img_tag = container.find('img')
                        if img_tag:
                            img = img_tag.get('src') or img_tag.get('data-src') or img_tag.get('data-original') or ""
                    if title and link and price: break
                    container = container.parent
                if title and link and price:
                    if link.startswith('/') and "sferis" in str(soup): link = "https://www.sferis.pl" + link
                    if "RAM" in title.upper() or "DDR5" in title.upper() or "PAMIĘĆ" in title.upper():
                        if "sferis" in link:
                            products.append({ 'title': title, 'price': price, 'link': link, 'image': img, 'shop': 'sferis', 'type': module_type })
            except Exception: continue
    unique = {}
    for p in products: unique[p['link']] = p
    return list(unique.values())

def parse_oleole(soup, module_type):
    products = []
    candidates = soup.select('.product-medium-box-intro__link, div[selenium-css="product-row"], div.product-row, div.product-box, div.product-medium-box')
    if not candidates:
        id_candidates = soup.find_all('a', id=re.compile(r'(ram|kingston|patriot|lexar|goodram|g-skill|corsair)', re.I))
        for a in id_candidates: candidates.append(a)
        candidates = list(set(candidates))

    for i, elem in enumerate(candidates):
        try:
            title_tag = None
            price_container = None
            if elem.name == 'a':
                title_tag = elem
                container = elem.find_parent('div', class_=re.compile(r'product-(box|row|medium-box)'))
                if not container: container = elem.parent.parent 
                price_container = container
            else:
                title_tag = elem.select_one('.product-medium-box-intro__link, h2.product-name a, .product-name a')
                price_container = elem

            if not title_tag: continue
            title = title_tag.get_text(strip=True)
            if not title and title_tag.has_attr('title'): title = title_tag['title']
            if not price_container: price_container = title_tag.parent
            
            price_text = ""
            sr_only_price = price_container.select_one('.parted-price .sr-only')
            if sr_only_price:
                price_text = sr_only_price.get_text(strip=True)
            else:
                parted_total = price_container.select_one('.parted-price-total')
                parted_decimal = price_container.select_one('.parted-price-decimal')
                if parted_total:
                    total_val = parted_total.get_text(strip=True)
                    decimal_val = parted_decimal.get_text(strip=True) if parted_decimal else "00"
                    price_text = f"{total_val},{decimal_val}"
                else:
                    price_tag = price_container.select_one('div[selenium-css="price-normal"], div.price-normal, .product-price, .price-value')
                    if price_tag: price_text = price_tag.get_text(" ", strip=True)
                    else: price_text = price_container.get_text(" ", strip=True)

            lower_price_text = price_text.lower()
            if ("rata" in lower_price_text or "x" in lower_price_text) and not sr_only_price and not parted_total:
                all_numbers = re.findall(r'(\d[\d\s\xa0]*[,.]?\d{0,2})', price_text)
                candidates_prices = []
                for num_str in all_numbers:
                    clean = re.sub(r'[^\d,]', '', num_str).replace(',', '.')
                    try:
                        val = float(clean)
                        if val > 100: candidates_prices.append(val)
                    except: pass
                if candidates_prices: price = max(candidates_prices) 
                else: continue 
            else:
                price_match = re.search(r'(\d[\d\s\xa0]*[,.]?\d{0,2})', price_text)
                if not price_match: continue
                original_number = price_match.group(1)
                has_separator = ',' in original_number or '.' in original_number
                clean_price = re.sub(r'[^\d,]', '', original_number).replace(',', '.')
                try: 
                    price = float(clean_price)
                    if price > 80000 and not has_separator: price = price / 100
                except: continue

            link = title_tag['href']
            if not link.startswith('http'): link = "https://www.oleole.pl" + link.strip()
            img_tag = price_container.select_one('img')
            img = img_tag.get('data-src') or img_tag.get('src') or "" if img_tag else ""

            if price >= MIN_PRICE_THRESHOLD:
                is_duplicate = False
                for p in products:
                    if p['link'] == link:
                        is_duplicate = True
                        break
                if not is_duplicate:
                    products.append({ 'title': title, 'price': price, 'link': link, 'image': img, 'shop': 'oleole', 'type': module_type })
        except Exception: continue
    return products

def parse_mediaexpert(soup, module_type):
    products = []
    candidates = soup.select('div.offer-box, div.c-offerBox, div.c-grid_item')
    
    for div in candidates:
        try:
            title_el = div.select_one('h2.name a, a.is-animate, .c-offerBox_title a, a.ui-link')
            if not title_el: continue
            title = title_el.get_text(strip=True)
            
            link = title_el['href']
            if not link.startswith('http'):
                link = 'https://www.mediaexpert.pl' + link
                
            price_text = ""
            main_price_div = div.select_one('.main-price, .price-box, .c-offerBox_price, .c-price_new')
            
            if main_price_div:
                whole_el = main_price_div.select_one('.whole')
                cents_el = main_price_div.select_one('.cents')
                if whole_el:
                    whole = whole_el.get_text(strip=True)
                    cents = cents_el.get_text(strip=True) if cents_el else "00"
                    price_text = f"{whole},{cents}"
                else:
                    price_text = main_price_div.get_text(" ", strip=True)
            if not price_text: continue

            price_match = re.search(r'(\d[\d\s\xa0]*[,.]\d{0,2})', price_text)
            if not price_match: continue
            clean_price = re.sub(r'[^\d,]', '', price_match.group(1)).replace(',', '.')
            try:
                price = float(clean_price)
                if price > 80000: price = price / 100
            except: continue
            
            img_tag = div.select_one('img')
            img = ""
            if img_tag:
                 img = img_tag.get('src') or img_tag.get('data-src') or ""

            if price >= MIN_PRICE_THRESHOLD:
                is_duplicate = False
                for p in products:
                    if p['link'] == link:
                        is_duplicate = True
                        break
                if not is_duplicate:
                    products.append({ 'title': title, 'price': price, 'link': link, 'image': img, 'shop': 'mediaexpert', 'type': module_type })
        except Exception: continue
    return products

def parse_komputronik(soup, module_type):
    products = []
    price_nodes = soup.select('div[data-price-type="final"], .price')
    
    for price_div in price_nodes:
        try:
            container = price_div.find_parent('li') or price_div.find_parent('div', class_=re.compile(r'(product-entry|pe2-item)'))
            if not container: container = price_div.parent.parent.parent
            if not container: continue

            title_tag = container.select_one('h2[class*="font-headline"], h2 a, .pe2-head a')
            if not title_tag: continue
            title = title_tag.get_text(strip=True)
            
            price_text = price_div.get_text(" ", strip=True)
            price_match = re.search(r'(\d[\d\s\xa0]*[,.]?\d{0,2})', price_text)
            if not price_match: continue
            clean_price = re.sub(r'[^\d,]', '', price_match.group(1)).replace(',', '.')
            try:
                price = float(clean_price)
                if price > 80000: price = price / 100
            except: continue
            
            link_tag = container.select_one('a[href*="/product/"], h2 a')
            if not link_tag and title_tag.name == 'a': link_tag = title_tag
            if link_tag:
                link = link_tag['href']
                if not link.startswith('http'):
                    link = "https://www.komputronik.pl" + link
            else: link = "#" 

            img_tag = container.select_one('img')
            img = img_tag.get('data-src') or img_tag.get('src') or "" if img_tag else ""

            if price >= MIN_PRICE_THRESHOLD:
                is_duplicate = False
                for p in products:
                    if p['link'] == link:
                        is_duplicate = True
                        break
                if not is_duplicate:
                    products.append({ 'title': title, 'price': price, 'link': link, 'image': img, 'shop': 'komputronik', 'type': module_type })
        except Exception: continue
    return products

PARSERS = {
    "parse_xkom": parse_xkom, "parse_morele": parse_morele, "parse_sferis": parse_sferis,
    "parse_oleole": parse_oleole, "parse_mediaexpert": parse_mediaexpert, "parse_komputronik": parse_komputronik
}

# ==========================================
# 4. SCENARIUSZE (ZADAŃ)
# ==========================================

SCENARIOS = [
    # --- X-KOM ---
    {
        "shop": "x-kom", "type": "single", 
        "base_url": "https://www.x-kom.pl/g-5/c/3402-pamieci-ram-ddr5.html",
        "params": "?sort_by=price_asc&f%5Bprice%5D%5Bfrom%5D=500&f1795-rodzaj-pamieci=293503-dimm&f1796-pojemnosc-calkowita=18899-pojedyncze-pamieci&f1797-taktowanie=228507-5600-mhz&f1797-taktowanie=292182-6000-6800-mhz&f1797-taktowanie=292183-od-7000-mhz&f1798-opoznienia-cycle-latency=223473-cl-40&f1798-opoznienia-cycle-latency=228503-cl-36&f1798-opoznienia-cycle-latency=255373-cl-32&f1798-opoznienia-cycle-latency=270216-cl-34&f1798-opoznienia-cycle-latency=288333-cl-30&f1798-opoznienia-cycle-latency=324658-cl-28&f1798-opoznienia-cycle-latency=332844-cl-26&f5973-liczba-modulow=109420-1&f1797-taktowanie=od-5600&f1797-taktowanie=do-9000&hide_unavailable=1",
        "page_param": "&page=", "cookie_texts": ["W porządku", "Akceptuję"], "parser_func": "parse_xkom", "max_pages": MAX_PAGES_PER_TASK
    },
    {
        "shop": "x-kom", "type": "dual", 
        "base_url": "https://www.x-kom.pl/g-5/c/3402-pamieci-ram-ddr5.html",
        "params": "?sort_by=price_asc&f%5Bprice%5D%5Bfrom%5D=500&f1795-rodzaj-pamieci=293503-dimm&f1796-pojemnosc-calkowita=54312-32-gb-2x16-gb&f1796-pojemnosc-calkowita=139451-64-gb-2x32-gb&f1796-pojemnosc-calkowita=282242-96-gb-2x48-gb&f1796-pojemnosc-calkowita=293628-48-gb-2x24-gb&f1797-taktowanie=228507-5600-mhz&f1797-taktowanie=292182-6000-6800-mhz&f1797-taktowanie=292183-od-7000-mhz&f1798-opoznienia-cycle-latency=223473-cl-40&f1798-opoznienia-cycle-latency=228503-cl-36&f1798-opoznienia-cycle-latency=255373-cl-32&f1798-opoznienia-cycle-latency=270216-cl-34&f1798-opoznienia-cycle-latency=288333-cl-30&f1798-opoznienia-cycle-latency=324658-cl-28&f1798-opoznienia-cycle-latency=332844-cl-26&f5973-liczba-modulow=109229-2&f1797-taktowanie=od-5600&f1797-taktowanie=do-9000&hide_unavailable=1",
        "page_param": "&page=", "cookie_texts": ["W porządku", "Akceptuję"], "parser_func": "parse_xkom", "max_pages": MAX_PAGES_PER_TASK
    },

    # --- MORELE ---
    {
        "shop": "morele", "type": "single",
        "base_url": "https://www.morele.net/kategoria/pamieci-ram-38/500,20000,,,,,,p,0,,,,7965O1919086,7971O454598.757416.740277.2161391.675242.2164899,7973O1945196.1951938.1933922.1933926.2069223.2135595.2154818.2135596.2154816.2139628.2154477.2166421.2175237.2178732.2235080.2235073,7974O395020.527381.744204.928725.985030.982636.980524.987639.1031908.1051232.1134877.1649868.1605104.2258001.2109647.2060845.1974756.2069226.1945198.1919087.1919089.1922851,21239O668024,sprzedawca:m/1/?noi",
        "strategy": "replace_pattern", "replace_from": "sprzedawca:m/1/", "replace_format": "sprzedawca:m/{}/",
        "cookie_texts": ["Zgadzam się", "Akceptuję"], "parser_func": "parse_morele", "max_pages": MAX_PAGES_PER_TASK
    },
    {
        "shop": "morele", "type": "dual",
        "base_url": "https://www.morele.net/kategoria/pamieci-ram-38/500,20000,,,,,,p,0,,,,7965O1919086,7971O454598.757416.740277.2161391.675242.2164899,7973O1945196.1951938.1933922.1933926.2069223.2135595.2154818.2135596.2154816.2139628.2154477.2166421.2175237.2178732.2235080.2235073,7974O395020.527381.744204.928725.985030.982636.980524.987639.1031908.1051232.1134877.1649868.1605104.2258001.2109647.2060845.1974756.2069226.1945198.1919087.1919089.1922851,21239O665142,sprzedawca:m/1/?noi",
        "strategy": "replace_pattern", "replace_from": "sprzedawca:m/1/", "replace_format": "sprzedawca:m/{}/",
        "cookie_texts": ["Zgadzam się", "Akceptuję"], "parser_func": "parse_morele", "max_pages": MAX_PAGES_PER_TASK
    },

    # --- SFERIS ---
    {
        "shop": "sferis", "type": "single",
        "base_url": "https://www.sferis.pl/pamieci-ram-2893",
        "params": "?f=price:500._,a818:13399.20165.55441.689166.13500.13542,a851:673744,a945:857247.741018.707568.727971.698603.86309.674482.673862.801441,a946:3213.751325.3253.754637.3296,a949:3284.25728.776437&l=60&o=price_asc",
        "page_param": "&p=", "cookie_texts": ["Akceptuję", "Zgadzam się"], "parser_func": "parse_sferis", "max_pages": MAX_PAGES_PER_TASK
    },
    {
        "shop": "sferis", "type": "dual",
        "base_url": "https://www.sferis.pl/pamieci-ram-2893",
        "params": "?f=price:500._,a818:13399.20165.55441.689166.13500.13542,a851:673744,a945:857247.741018.707568.727971.698603.86309.674482.673862.801441,a946:3213.3253.3296.751325.754637,a949:3323.555576.751326.25728.776437&l=60&o=price_asc",
        "page_param": "&p=", "cookie_texts": ["Akceptuję", "Zgadzam się"], "parser_func": "parse_sferis", "max_pages": MAX_PAGES_PER_TASK
    },

    # --- OLEOLE ---
    {
        "shop": "oleole", "type": "single",
        "base_url": "https://www.oleole.pl/pamieci-ram,a1,pojemnosc-pamieci!16-gb:32-gb:64-gb:12:14:18,v37:38:41:44:52:53:54:55:56:59:60,i10,m23,aa18:19:20:21:22:23:24:26:29,od500,d3.bhtml",
        "strategy": "replace_pattern", "replace_from": "d3.bhtml", "replace_format": "d3,strona-{}.bhtml",
        "cookie_texts": ["Akceptuj", "Zamknij"], "parser_func": "parse_oleole", "max_pages": MAX_PAGES_PER_TASK
    },
    {
        "shop": "oleole", "type": "dual",
        "base_url": "https://www.oleole.pl/pamieci-ram,a1,pojemnosc-pamieci!16-gb:32-gb:64-gb:12:14:18,v37:38:41:44:52:53:54:55:56:59:60,i10,m24,aa18:19:20:21:22:23:24:26:29,od500,d3.bhtml",
        "strategy": "replace_pattern", "replace_from": "d3.bhtml", "replace_format": "d3,strona-{}.bhtml",
        "cookie_texts": ["Akceptuj", "Zamknij"], "parser_func": "parse_oleole", "max_pages": MAX_PAGES_PER_TASK
    },

    # --- MEDIA EXPERT ---
    {
        "shop": "mediaexpert", "type": "single",
        "base_url": "https://www.mediaexpert.pl/komputery-i-tablety/podzespoly-komputerowe/pamieci-ram/typ-pamieci_ddr-5/ilosc-modulow_1/czestotliwosc-pracy-mhz_6000.5600.6400.6200.7000.7200.6800/pojemnosc-gb_16.32.64.48.96.24/zastosowanie_komputer-pc",
        "params": "?availability_name=Dost%C4%99pny&limit=50&sort=price_asc",
        "page_param": "&page=", "cookie_texts": ["Akceptuj", "Zaakceptuj"], "parser_func": "parse_mediaexpert", "max_pages": MAX_PAGES_PER_TASK
    },
    {
        "shop": "mediaexpert", "type": "dual",
        "base_url": "https://www.mediaexpert.pl/komputery-i-tablety/podzespoly-komputerowe/pamieci-ram/typ-pamieci_ddr-5/ilosc-modulow_2/czestotliwosc-pracy-mhz_6000.5600.6400.6200.7000.7200.6800/pojemnosc-gb_16.32.64.48.96.24/zastosowanie_komputer-pc",
        "params": "?availability_name=Dost%C4%99pny&limit=50&sort=price_asc",
        "page_param": "&page=", "cookie_texts": ["Akceptuj", "Zaakceptuj"], "parser_func": "parse_mediaexpert", "max_pages": MAX_PAGES_PER_TASK
    },

    # --- KOMPUTRONIK ---
    {
        "shop": "komputronik", "type": "single",
        "base_url": "https://www.komputronik.pl/category/437/pamiec-ram.html",
        "params": "?a[1069][]=5600&a[1069][]=9000&a[112422][]=88133&a[112423][]=159963&a[112423][]=88129&a[112423][]=88130&a[112423][]=88132&a[112423][]=92023&a[114058][]=134619&a[289][]=105417&a[290][]=16&a[290][]=64&a[294][]=9&a[294][]=42&alt=1&by=f_price_10&pr10[]=500&pr10[]=500000&showBuyActiveOnly=1&sort=1&filter=1",
        "page_param": "&p=", "cookie_texts": ["Akceptuję", "Zgadzam się"], "parser_func": "parse_komputronik", "max_pages": MAX_PAGES_PER_TASK
    },
    {
        "shop": "komputronik", "type": "dual",
        "base_url": "https://www.komputronik.pl/category/437/pamiec-ram.html",
        "params": "?a[1069][]=5600&a[1069][]=9000&a[112422][]=88134&a[112423][]=159963&a[112423][]=88129&a[112423][]=88130&a[112423][]=88132&a[112423][]=92023&a[114058][]=134619&a[289][]=105417&a[290][]=16&a[290][]=64&a[294][]=9&a[294][]=42&alt=1&by=f_price_10&pr10[]=500&pr10[]=500000&showBuyActiveOnly=1&sort=1&filter=1",
        "page_param": "&p=", "cookie_texts": ["Akceptuję", "Zgadzam się"], "parser_func": "parse_komputronik", "max_pages": MAX_PAGES_PER_TASK
    }
]

# ==========================================
# 5. GŁÓWNA PĘTLA ASYNC - RÓWNOLEGŁA
# ==========================================

async def process_shop_tasks(context, shop_name, tasks):
    print(f"🚀 [START] Wątek dla sklepu: {shop_name.upper()}...")
    shop_products = []
    
    # Tworzymy nową kartę (page) w ramach wspólnego kontekstu (jednego okna)
    page = await context.new_page()

    try:
        for task in tasks:
            group_type = task['type']
            seen_page_hashes = set()
            links = []
            if task.get('strategy') == 'replace_pattern':
                 for i in range(1, task['max_pages'] + 1):
                    if i == 1: links.append(task['base_url'])
                    else:
                        pat = task['replace_format'].format(i)
                        links.append(task['base_url'].replace(task['replace_from'], pat))
            else:
                for i in range(1, task['max_pages'] + 1):
                    if i == 1: links.append(task['base_url'] + task['params'])
                    else: links.append(task['base_url'] + task['params'] + task['page_param'] + str(i))
            
            for i, url in enumerate(links, 1):
                try:
                    await page.goto(url, wait_until="domcontentloaded", timeout=45000)
                    await handle_cookies(page, task['cookie_texts'])
                    
                    if shop_name in ['oleole', 'mediaexpert', 'komputronik']:
                        try: await page.wait_for_load_state("networkidle", timeout=5000)
                        except: pass
                        
                    await human_scroll(page)
                    content = await page.content()
                    
                    parser_name = task['parser_func']
                    parser = PARSERS.get(parser_name)
                    
                    if parser:
                        items = parser(BeautifulSoup(content, 'html.parser'), group_type)
                        
                        if not items:
                            break
                        
                        # --- SMART STOP: DUPLICATE CHECK ---
                        first = items[0]
                        current_hash = f"{first['title']}{first['price']}"
                        if current_hash in seen_page_hashes:
                            break
                        seen_page_hashes.add(current_hash)

                        # --- SMART STOP: KOMPUTRONIK ---
                        if shop_name == 'komputronik' and len(items) < 20:
                             shop_products.extend(items)
                             break
                        
                        shop_products.extend(items)
                        print(f"   + {shop_name.upper()} [{group_type}]: strona {i} -> {len(items)} szt.")
                    
                    # --- OPÓŹNIENIE ZGODNE Z TARGETED ---
                    await asyncio.sleep(random.randint(3, 6))
                    
                except Exception as e:
                    print(f"   !!! Błąd {shop_name} strona {i}: {e}")
                    break
    
    except Exception as e:
        print(f"!!! KRYTYCZNY BŁĄD WĄTKU {shop_name}: {e}")
    finally:
        await page.close() 
        print(f"🏁 [KONIEC] Wątek {shop_name.upper()} zakończony. Pobrano łącznie: {len(shop_products)}")
    
    return shop_products

# ==========================================
# 6. LOGIKA EXPORTU I RAPORTOWANIA
# ==========================================

def export_to_real_excel(all_products):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("!!! BRAK BIBLIOTEKI OPENPYXL !!!")
        return

    try:
        all_products.sort(key=lambda x: (x['type'], x['shop'], x['price']))
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Wszystkie Produkty"
        headers = ['Typ (Single/Dual)', 'Sklep', 'Tytuł', 'Cena (PLN)', 'Link', 'Data pobrania']
        ws.append(headers)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        scan_date = datetime.now().strftime("%Y-%m-%d %H:%M")
        for p in all_products:
            ws.append([
                p['type'].upper(),
                p['shop'].upper(),
                p['title'],
                p['price'],
                p['link'],
                scan_date
            ])
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 60
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 50
        ws.column_dimensions['F'].width = 20
        for row in range(2, len(all_products) + 2):
            cell = ws.cell(row=row, column=4)
            cell.number_format = '#,##0.00 "zł"'
        try: wb.save(DEBUG_XLSX_FILE)
        except PermissionError: pass
    except Exception:
        print(f"Błąd generowania XLSX: {Exception}")

def export_all_products_to_excel_csv(all_products):
    try:
        all_products.sort(key=lambda x: (x['type'], x['shop'], x['price']))
        with open(DEBUG_CSV_FILE, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f, delimiter=';')
            writer.writerow(['Typ', 'Sklep', 'Tytuł', 'Cena (PLN)', 'Link'])
            for p in all_products:
                price_str = f"{p['price']:.2f}".replace('.', ',')
                writer.writerow([p['type'].upper(), p['shop'].upper(), p['title'], price_str, p['link']])
        print(f"--> [DEBUG] Zaktualizowano plik CSV: {DEBUG_CSV_FILE}")
    except Exception:
        print(f"Błąd generowania Excel Debug CSV: {Exception}")

def export_to_csv(categorized_data):
    try:
        with open(CSV_FILE, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f, delimiter=';') 
            writer.writerow(['Kategoria', 'Sklep', 'Produkt', 'Cena', 'Link', 'Data'])
            for cat_name, shops in categorized_data.items():
                for shop_name, product in shops.items():
                    writer.writerow([cat_name, shop_name.upper(), product['title'], f"{product['price']:.2f}".replace('.', ','), product['link'], datetime.now().strftime("%Y-%m-%d %H:%M")])
    except Exception: pass

def categorize_and_pick_best(all_products):
    CAT_MAP = {
        ("single", 16): "16GB (1x16GB)", ("single", 32): "32GB (1x32GB)",
        ("single", 48): "48GB (1x48GB)", ("single", 64): "64GB (1x64GB)",
        ("dual", 32): "32GB (2x16GB)", ("dual", 48): "48GB (2x24GB)",
        ("dual", 64): "64GB (2x32GB)", ("dual", 96): "96GB (2x48GB)"
    }
    try:
        with open(DEBUG_JSON_FILE, "w", encoding="utf-8") as f:
            json.dump(all_products, f, indent=2, ensure_ascii=False)
        print(f"--> [DEBUG] Zapisano surowe dane JSON: {DEBUG_JSON_FILE}")
    except Exception as e:
        print(f"Błąd zapisu JSON: {e}")

    results = {name: {} for name in CAT_MAP.values()}
    results["Najtańsza DDR5 (Ogólnie)"] = {}
    sorted_products = sorted(all_products, key=lambda x: x['price'])
    
    for product in sorted_products:
        shop = product['shop']
        p_type = product['type']
        title = product['title'].upper()
        gb_matches = re.findall(r'(\d+)\s?GB', title)
        capacity = 0
        if gb_matches:
            caps = [int(x) for x in gb_matches]
            valid_caps = [16, 32, 48, 64, 96]
            found_caps = [c for c in caps if c in valid_caps]
            if found_caps: capacity = max(found_caps)
        if capacity == 0: continue
        if p_type == 'dual':
             if capacity == 16: capacity = 32
             elif capacity == 24: capacity = 48
             elif capacity == 32 and "2X32" in title: capacity = 64 
        cat_key = (p_type, capacity)
        if cat_key in CAT_MAP:
            cat_name = CAT_MAP[cat_key]
            if shop not in results[cat_name]:
                results[cat_name][shop] = product
        if shop not in results["Najtańsza DDR5 (Ogólnie)"]:
            results["Najtańsza DDR5 (Ogólnie)"][shop] = product
    return results

def update_history(best_deals):
    history = []
    if os.path.exists(HISTORY_FILE):
        try:
            with open(HISTORY_FILE, 'r') as f: history = json.load(f)
        except: pass
    gen_cat = best_deals.get("Najtańsza DDR5 (Ogólnie)", {})
    prices = [p['price'] for p in gen_cat.values()]
    if prices:
        min_price = min(prices)
        history.append({"timestamp": datetime.now().strftime("%Y-%m-%d %H:%M"), "price": min_price})
        history = history[-50:]
        with open(HISTORY_FILE, 'w') as f: json.dump(history, f)
    return history

def generate_pro_html(categorized_data, history_data, next_scan_time):
    labels = []
    for h in history_data:
        try:
            dt_obj = datetime.strptime(h['timestamp'], "%Y-%m-%d %H:%M")
            labels.append(dt_obj.strftime("%d.%m %H:%M"))
        except: labels.append(h['timestamp'])
    data_points = [h['price'] for h in history_data]
    
    cards_html = ""
    for cat_name, shops_data in categorized_data.items():
        rows_html = ""
        if not shops_data:
             rows_html = "<div class='text-center text-gray-500 py-4'>Brak ofert</div>"
        else:
            sorted_shops = sorted(shops_data.values(), key=lambda x: x['price'])
            best_price = sorted_shops[0]['price']
            for i, product in enumerate(sorted_shops):
                is_winner = (i == 0)
                wrapper_class = "bg-gray-700/50 border-green-500/50" if is_winner else "bg-gray-800 border-transparent"
                text_color = "text-green-400" if is_winner else "text-gray-400"
                badges_html = '<span class="text-[10px] bg-green-900/60 text-green-300 px-2 py-0.5 rounded border border-green-700/50">NAJTANIEJ</span>' if is_winner else ''
                safe_title = html.escape(product['title'])
                safe_shop = html.escape(product['shop'])
                rows_html += f"""
                <div class="flex items-start justify-between p-3 {wrapper_class} border-l-4 mb-2 rounded transition-colors duration-200">
                    <div class="flex-1 min-w-0 mr-3">
                        <div class="flex items-center gap-2 mb-1">
                            <span class="text-[10px] font-bold px-2 py-0.5 rounded bg-gray-900 text-gray-300 uppercase tracking-wider">{safe_shop}</span>
                            {badges_html}
                        </div>
                        <div class="text-sm text-gray-300 leading-tight" title="{safe_title}">{safe_title}</div>
                    </div>
                    <div class="text-right flex flex-col items-end min-w-[80px]">
                        <div class="text-lg font-bold {text_color} font-mono">{product['price']:.2f} zł</div>
                        <a href="{product['link']}" target="_blank" class="mt-2 text-[10px] px-3 py-1 rounded-full bg-gray-700 hover:bg-gray-600 text-gray-300 transition-colors uppercase font-bold tracking-wide">IDŹ DO SKLEPU</a>
                    </div>
                </div>"""
        card = f"""
        <div class="bg-gray-800 rounded-xl overflow-hidden border border-gray-700 shadow-lg flex flex-col">
            <div class="bg-gray-900/80 px-4 py-3 border-b border-gray-700 flex justify-between items-center backdrop-blur-sm">
                <h3 class="text-gray-100 font-bold text-sm uppercase tracking-wider truncate" title="{cat_name}">{cat_name}</h3>
            </div>
            <div class="p-3 flex-1 overflow-y-auto max-h-[300px] scrollbar-thin scrollbar-thumb-gray-700 scrollbar-track-transparent">
                {rows_html}
            </div>
        </div>"""
        cards_html += card

    html_content = f"""<!DOCTYPE html><html lang="pl" class="dark"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0"><title>RAM Sniper PARALLEL</title>
    <!-- iOS Viewport & Scaling Fix -->
    <meta name="apple-mobile-web-app-capable" content="yes">
    <meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
    <script src="https://cdn.tailwindcss.com"></script><script src="https://cdn.jsdelivr.net/npm/chart.js"></script><script>tailwind.config = {{ darkMode: 'class', theme: {{ extend: {{ colors: {{ gray: {{ 900: '#111827', 800: '#1f2937', 700: '#374151' }} }} }} }} }}</script>
    <style>
        body {{ background-color: #0f1115; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Helvetica, Arial, sans-serif; -webkit-text-size-adjust: 100%; }}
    </style>
    </head><body class="text-gray-300 min-h-screen flex flex-col"><nav class="sticky top-0 z-50 bg-gray-900/90 backdrop-blur border-b border-gray-800"><div class="max-w-7xl mx-auto px-4 py-4 flex justify-between items-center"><div class="flex items-center gap-3"><div class="bg-green-600 p-2 rounded-lg text-white"><svg class="w-6 h-6" fill="none" stroke="currentColor" viewBox="0 0 24 24"><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M13 10V3L4 14h7v7l9-11h-7z"></path></svg></div><div><h1 class="text-xl font-bold text-white tracking-tight">RAM Sniper <span class="text-green-500">PARALLEL</span></h1><div class="flex items-center gap-2 text-xs text-green-400 font-mono"><span class="relative flex h-2 w-2"><span class="animate-ping absolute inline-flex h-full w-full rounded-full bg-green-400 opacity-75"></span><span class="relative inline-flex rounded-full h-2 w-2 bg-green-500"></span></span>X-Kom + Morele + Sferis + OleOle + MediaExpert + Komputronik</div></div></div></div></nav><main class="flex-1 max-w-7xl w-full mx-auto px-4 py-8 space-y-8"><div class="grid grid-cols-1 md:grid-cols-3 gap-6"><div class="md:col-span-2 bg-gray-800 rounded-xl p-6 border border-gray-700"><h3 class="text-gray-400 text-sm font-medium mb-4">Trend Cenowy</h3><div class="h-48 relative w-full"><canvas id="priceChart"></canvas></div></div><div class="bg-gray-800 rounded-xl p-6 border border-gray-700 flex flex-col justify-center items-center text-center"><div class="mb-2 p-3 bg-gray-900 rounded-full text-green-500"><svg class="w-8 h-8" fill="none" stroke="currentColor" viewBox="0 0 24 24"><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M12 8v4l3 3m6-3a9 9 0 11-18 0 9 9 0 0118 0z"></path></svg></div><h3 class="text-gray-400 text-sm">Ostatnia aktualizacja</h3><p class="text-2xl font-bold text-white mt-1">{datetime.now().strftime("%H:%M")}</p></div></div><div><h2 class="text-xl font-bold text-white mb-4 flex items-center gap-2"><span class="w-1 h-6 bg-green-600 rounded-full"></span>Wyniki Celowane (Równoległe)</h2><div class="grid grid-cols-1 md:grid-cols-2 xl:grid-cols-3 gap-6">{cards_html}</div></div></main><script>const ctx = document.getElementById('priceChart').getContext('2d');new Chart(ctx, {{ type: 'line', data: {{ labels: {json.dumps(labels)}, datasets: [{{ label: 'Cena', data: {json.dumps(data_points)}, borderColor: '#22c55e', backgroundColor: 'rgba(34, 197, 94, 0.1)', borderWidth: 2, tension: 0.4, fill: true, pointRadius: 3 }}] }}, options: {{ responsive: true, maintainAspectRatio: false, plugins: {{ legend: {{ display: false }} }}, scales: {{ y: {{ grid: {{ color: '#374151' }}, ticks: {{ color: '#9ca3af' }} }}, x: {{ grid: {{ display: false }}, ticks: {{ color: '#9ca3af' }} }} }} }} }});</script></body></html>"""
    with open(REPORT_FILE, "w", encoding="utf-8") as f: f.write(html_content)
    print(f"Raport zaktualizowany: {REPORT_FILE}")

async def main():
    print("--- RAM Sniper PARALLEL (6 Shops Concurrent) ---")
    if not os.path.exists(USER_DATA_DIR): os.makedirs(USER_DATA_DIR)
    
    async with async_playwright() as p:
        # Uruchamiamy przeglądarkę RAZ (jedno okno)
        context = await p.chromium.launch_persistent_context(
            user_data_dir=USER_DATA_DIR,
            headless=False,
            args=["--disable-blink-features=AutomationControlled", "--start-maximized"],
            viewport=None
        )
        
        await asyncio.sleep(2)
        
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Uruchamianie workerów...")
        
        tasks_by_shop = {}
        for task in SCENARIOS:
            shop = task['shop']
            if shop not in tasks_by_shop:
                tasks_by_shop[shop] = []
            tasks_by_shop[shop].append(task)
            
        async_tasks = []
        for shop_name, tasks in tasks_by_shop.items():
            async_tasks.append(process_shop_tasks(context, shop_name, tasks))
            
        results_list = await asyncio.gather(*async_tasks, return_exceptions=True)
        
        all_found_products = []
        for res in results_list:
            if isinstance(res, list):
                all_found_products.extend(res)
            else:
                print(f"!!! Błąd jednego z workerów: {res}")

        await context.close()
        
        if all_found_products:
            print(f"\n--- SKANOWANIE ZAKOŃCZONE. Łącznie produktów: {len(all_found_products)} ---")
            print("Generowanie raportów...")
            
            export_all_products_to_excel_csv(all_found_products)
            export_to_real_excel(all_found_products)

            best_deals = categorize_and_pick_best(all_found_products)
            history = update_history(best_deals)
            generate_pro_html(best_deals, history, "DONE")
            export_to_csv(best_deals)
            
            print(f"SUKCES! Raporty gotowe: {REPORT_FILE}")
        else:
            print("!!! Pusta lista produktów. Coś poszło nie tak.")

if __name__ == "__main__":
    asyncio.run(main())